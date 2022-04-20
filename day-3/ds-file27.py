# reading an excel file of multiple sheets

#importing the load_workbook class from openpyxl
from openpyxl import load_workbook

#selecting the excel file
x=load_workbook('h_excel3.xlsx')

#selecting the sheetnames
y=x.sheetnames


#select each sheet and make it active
for m in y:
    p=x[m]    #will make the sheet active
    print ('Sheet :: %s' % (m))
    for i in range(1,p.max_row+1):
        for j in range(1,p.max_column+1):
            print ('%-10s' % p.cell(i,j).value, ' ', end='')
        print ()
x.close()

            
        
